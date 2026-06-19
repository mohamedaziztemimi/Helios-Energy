"""
Excel generation tool. Produces a real .xlsx file using openpyxl — not
mocked. Unlike the PDF/Word tools, Excel's natural shape is tabular, so
this tool's input is just sheet name + rows, no separate "content" field.
"""

import os
import re
import uuid

from langchain_core.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _safe_filename(title: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9_-]+", "_", title.strip()).strip("_").lower()
    return f"{slug or 'report'}_{uuid.uuid4().hex[:8]}.xlsx"


def make_generate_excel_tool(output_dir: str):
    """
    output_dir: the directory this run's generated files should be written
    to, e.g. /tmp/agent_outputs/<run_id>/.
    """

    @tool
    def generate_excel(title: str, sheet_name: str, rows: list[list[str]]) -> str:
        """
        Generate a real, downloadable Excel (.xlsx) workbook. Returns the file path.

        title: used to name the file (not written inside the spreadsheet).
        sheet_name: the name of the sheet tab, e.g. "Plant Summary".
        rows: a list of rows, where the first row is the header.
            Example: [["Plant", "Energy (kWh)"], ["Plant C1-001", 1234.5]]
        """
        os.makedirs(output_dir, exist_ok=True)
        filename = _safe_filename(title)
        filepath = os.path.join(output_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31] if sheet_name else "Sheet1"  # Excel sheet name limit

        for row in rows:
            ws.append(row)

        if rows:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")

        for col_cells in ws.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(length + 4, 40)

        wb.save(filepath)
        return filepath

    return generate_excel